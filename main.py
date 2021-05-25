import pandas as pd
import collections
import openpyxl

'''
PAYERACCOUNTNUMBER - номер счета плательщика
PAYEEACCOUNTNUMBER - номер счета получателя
AMOUNT - сумма операции
DATEIN - дата операции
'''

header_name = ['PAYERACCOUNTNUMBER', 'PAYEEACCOUNTNUMBER', 'AMOUNT', 'DATEIN']

loaded_table = pd.read_csv('Test_data.csv',sep=';') # читаем таблицу, с помощью параметра sep делим таблицу.

loaded_dict_PAYER = loaded_table['PAYERACCOUNTNUMBER'].to_dict()#преобразуем столбец  PAYERACCOUNTNUMBER в словарь
loaded_dict_PAYEE = loaded_table['PAYEEACCOUNTNUMBER'].to_dict()#преобразуем столбец PAYEEACCOUNTNUMBER в словарь
loaded_dict_AMOUNT = loaded_table['AMOUNT'].to_dict()#преобразуем столбец AMOUNT в словарь
loaded_dict_DATEIN = loaded_table['DATEIN'].to_dict()#преобразуем столбец DATEIN в словарь

def write_header_excel(list_header,lenth,name):
    wb = openpyxl.load_workbook(filename=name+'.xlsx')
    names = name.split(' ')
    ws = wb['Factor_Table']
    if lenth == 1:
        if 'Amount' in names and 'Payer' in names:
            for v in range(0, 3, 2):  # запись в первый столбец наименований дальнейших данных
                if v == 0:
                    ws.cell(row=1, column=v+1).value = list_header[v]
                else:
                    ws.cell(row=1, column=v).value = list_header[v]
        elif 'Amount' in names and 'Payee' in names:
            for v in range(1,3):
                    ws.cell(row=1, column=v).value = list_header[v]
        elif 'Operation' in names and 'Payer' in names:
            for v in range(0, 3, 2):  # запись в первый столбец наименований дальнейших данных
                if v == 0:
                    ws.cell(row=1, column=v+1).value = list_header[v]
                else:
                    ws.cell(row=1, column=v).value = 'Total operation'
        elif 'Operation' in names and 'Payee' in names:
            for v in range(1,3):
                    ws.cell(row=1, column=v).value = list_header[v]
        elif 'Operation'  in names and 'transactions' in names:
            for count in range(1,4):
                if count == 3:
                    ws.cell(row=1, column=count).value = 'Total Operation'
                else:
                    ws.cell(row=1, column=count).value = list_header[count]
        elif 'Amount' in names and 'transactions' in names:
            for count in range(1, 4):
                ws.cell(row=1, column=count).value = list_header[count]



    wb.save(filename=name+'.xlsx')
    wb.close()

#Функция на подсчёт общий суммы операций
def total_amount(dicti,array,name):
    total_array = collections.defaultdict(dict)
    total_array.clear()
    for count in range(len(dicti)):
        for number in array:
            if dicti[count] == number:
                total_array.setdefault(number, []).append(float(loaded_dict_AMOUNT[count].replace(',', '.')))
    for key in total_array:
        total_array[key] = str(sum(total_array[key]))

    return write_excel(total_array,name)

def max_amount(dicti,array,name):
    total_array_max = collections.defaultdict(dict)
    for count in range(len(dicti)):
        for number in array:
            if dicti[count] == number:
                total_array_max.setdefault(number, []).append(float(loaded_dict_AMOUNT[count].replace(',', '.')))
    for key in total_array_max:
        total_array_max[key] = str(max(total_array_max[key]))
    return write_excel(total_array_max, name)

#Функиця на подсчёт	общее количество операций
def total_operati(dicti,array,name):
    total_operatin = collections.defaultdict(dict)
    for number in array:
        total_operatin[number] = 0

    for count in range(len(dicti)):
        for number in array:
            if dicti[count] == number:
                total_operatin[number] = total_operatin[number] + 1

    return write_excel(total_operatin,name)

def write_excel(array,name):
    table_taP = pd.DataFrame.from_dict(array, orient='index')
    size_df = table_taP.shape
    writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
    table_taP.to_excel(writer, sheet_name='Factor_Table')
    writer.save()
    # return (write_header_excel(header_name,size_df[1],name))

def wr_excel_multi(df,name):
    size_df = df.shape
    writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Factor_Table')
    writer.save()
    # return (write_header_excel(header_name, size_df[1], name))

def transpon(array,name):  # функция преобразования из словаря в табличку
    df = pd.DataFrame.from_dict(array,orient='columns')  # делаем из словаря таблицу
    index = pd.MultiIndex.from_frame(df)  # делаем мультииндекс таблицу
    new_df = pd.DataFrame(df.unstack())# создаем таблицу с мультииндексами
    new_df = new_df.dropna()
    return wr_excel_multi(new_df,name)

def tr(array,name):
    df = pd.DataFrame.from_dict(array, orient='columns')  # делаем из словаря таблицу
    index = pd.MultiIndex.from_frame(df)  # делаем мультииндекс таблицу
    new_df = pd.DataFrame(df.unstack())  # создаем таблицу с мультииндексами
    print(new_df)
    # new_df = new_df.dropna()
    # return wr_excel_multi(new_df, name)
PAYER = []#массив счетов плательщика
PAYEE = []#массив счетов получателя
DATEIN=[]#массив для дат

tA_PAYER_PAYEE = collections.defaultdict(dict) #Общая сумма операций от плательщика x получателю y (для каждой пары плательщик-получатель)
tO_PAYER_PAYEE = collections.defaultdict(dict) #Общая количество операций от плательщика x получателю y (для каждой пары плательщик-получатель)

tA_PAYER_PAYEE_DATEIN = collections.defaultdict(dict)#Общая сумма операций от плательщика x получателю y за дату z (для каждой пары плательщик-получатель и каждой даты)
tO_PAYER_PAYEE_DATEIN = collections.defaultdict(dict)#Общая количество операций от плательщика x получателю y за дату z (для каждой пары плательщик-получатель и каждой даты)

#цикл на заполнение массивов
for count in range(len(loaded_dict_PAYER)):
    PAYER.append(loaded_dict_PAYER[count])
    PAYEE.append(loaded_dict_PAYEE[count])
    DATEIN.append(loaded_dict_DATEIN[count])

    tA_PAYER_PAYEE[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] = 0.0

    tA_PAYER_PAYEE_DATEIN[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] = {}
    tA_PAYER_PAYEE_DATEIN[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]][loaded_dict_DATEIN[count]] = 0.0

    tO_PAYER_PAYEE[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] = 0

    tO_PAYER_PAYEE_DATEIN[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] = collections.defaultdict(dict)
    tO_PAYER_PAYEE_DATEIN[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]][loaded_dict_DATEIN[count]] = 0

#цикл на подсчёт общей суммы операций от плательщика x получателю y (для каждой пары плательщик-получатель)
for count in range(len(loaded_dict_PAYER)):
    if loaded_dict_PAYER[count] in tA_PAYER_PAYEE and (loaded_dict_PAYEE[count]) in tA_PAYER_PAYEE[loaded_dict_PAYER[count]]:
        tA_PAYER_PAYEE[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] = tA_PAYER_PAYEE[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] + float(loaded_dict_AMOUNT[count].replace(',', '.'))

#цикл подсчёта Общая сумма операций от плательщика x получателю y за дату z (для каждой пары плательщик-получатель и каждой даты)
for count in range(len(loaded_dict_PAYER)):
    payer = loaded_dict_PAYER[count]
    payee = loaded_dict_PAYEE[count]
    datein = loaded_dict_DATEIN[count]
    if payer in tA_PAYER_PAYEE_DATEIN and payee in tA_PAYER_PAYEE_DATEIN[payer] and  datein in tA_PAYER_PAYEE_DATEIN[payer][payee]:
        tA_PAYER_PAYEE_DATEIN[payer][payee][datein] = tA_PAYER_PAYEE_DATEIN[payer][payee][datein] + float(loaded_dict_AMOUNT[count].replace(',', '.'))

#цикл на подсчётколичество операций от плательщика x получателю y (для каждой пары плательщик-получатель)
for count in range(len(loaded_dict_PAYER)):
    if loaded_dict_PAYER[count] in tO_PAYER_PAYEE and (loaded_dict_PAYEE[count]) in tO_PAYER_PAYEE[loaded_dict_PAYER[count]]:
            tO_PAYER_PAYEE[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] = tO_PAYER_PAYEE[loaded_dict_PAYER[count]][loaded_dict_PAYEE[count]] + 1

#цикл подсчёта общего количества операций от плательщика x получателю y за дату z (для каждой пары плательщик-получатель и каждой даты)
for count in range(len(loaded_dict_PAYER)):
    payer = loaded_dict_PAYER[count]
    payee = loaded_dict_PAYEE[count]
    datein = loaded_dict_DATEIN[count]
    if payer in tO_PAYER_PAYEE_DATEIN and payee in tO_PAYER_PAYEE_DATEIN[payer] and  datein in tO_PAYER_PAYEE_DATEIN[payer][payee]:
        tO_PAYER_PAYEE_DATEIN[payer][payee][datein] = tO_PAYER_PAYEE_DATEIN[payer][payee][datein] + 1

PAYER = set(PAYER)# с помощью функции set убираем дубликаты счетов плательщика
PAYEE = set(PAYEE)# с помощью функции set убираем дубликаты счетов получателя
DATEIN = set(DATEIN)# с помощью функции set убираем дубликаты дат

# print(tA_PAYER_PAYEE_DATEIN)
# print(tO_PAYER_PAYEE_DATEIN)

total_amount_PAYER = total_amount(loaded_dict_PAYER,PAYER,'a)Total Amount Payer') #общая сумма операций плательщика
total_operation_PAYER = total_operati(loaded_dict_PAYER,PAYER,'c)Total Operation Payer')# общие количество операций плательщика
max_amount_operation_PAYER = max_amount(loaded_dict_PAYER,PAYER,'k)Max Amount Payer')

total_amount_PAYEE = total_amount(loaded_dict_PAYEE,PAYEE,'b)Total Amount Payee')#общая сумма операций получателя
total_operation_PAYEE = total_operati(loaded_dict_PAYEE,PAYEE,'d)Total Operation Payee')#общая количество операций получателя
max_amount_operation_PAYEE = max_amount(loaded_dict_PAYEE,PAYEE,'l)Max Amount Payee')

total_amount_DATEIN = total_amount(loaded_dict_DATEIN,DATEIN,'i)Total Amount DateIn')#общая сумма операций за каждую дату
total_operation_DATEIN = total_operati(loaded_dict_DATEIN,DATEIN,'j)Total Operation DateIn')#общая количество операций за дату

tA_par_pae = transpon(tA_PAYER_PAYEE,'e)Total Amount of transactions from Payer to Payee')#Общая сумма операций от плательщика x получателю y (для каждой пары плательщик-получатель)
tO_par_pae = transpon(tO_PAYER_PAYEE,'f)Total Operation of transactions from Payer to Payee')#Общee количество операций от плательщика x получателю y (для каждой пары плательщик-получатель)

# ta_par_pae_date = tr(tA_PAYER_PAYEE_DATEIN,'g)Total Amount of transactions from Payer to Payee for date')
# print(ta_par_pae_date)

print('Всего плательщиков: '+str(len(PAYER))+'\n'+'Всего получателей: ' + str(len(PAYEE)))